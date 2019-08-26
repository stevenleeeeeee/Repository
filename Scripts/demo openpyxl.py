"""
官方文档：https://openpyxl.readthedocs.io/en/stable/charts/bar.html
"""

import datetime
from random import choice
from time import time
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

addr = "openpyxl.xlsx"                  # 设置文件
wb = load_workbook(addr)                # 打开文件
ws = wb.create_sheet()                  # 创建一张新表
ws.append(['TIME', 'TITLE', 'A-Z'])     # 第一行输入

# 输入内容（500行数据）
for i in range(500):
    TIME = datetime.datetime.now().strftime("%H:%M:%S")     # 创建数据
    TITLE = str(time())                                     # 
    A_Z = get_column_letter(choice(range(1, 50)))           # 
    ws.append([TIME, TITLE, A_Z])                           # 追加输入行，每个元素各占一列


row_max = ws.max_row                    # 获取最大行
con_max = ws.max_column                 # 获取最大列

# 把上面写入内容打印在控制台
for j in ws.rows:	                    # ws.rows 获取每一行数据
    for n in j:
        print(n.value, end="\t")        # n.value 获取单元格的值
    print()

# 保存，save（必须要写文件名（绝对地址）默认 py 同级目录下，只支持 xlsx 格式）
wb.save(addr)


# -------------------------------------------------

# 打开已有
from openpyxl  import load_workbook
wb2 = load_workbook('文件名称.xlsx')     # 载入文件。 wb = Workbook()  ----> 创建文件对象

# 获取第一个sheet
# ws = wb.active     

# 创建表 sheet
ws1 = wb.create_sheet("Mysheet")        # 方式一：插入到最后(default)
ws2 = wb.create_sheet("Mysheet",0)      # 方式二：插入到最开始的位置
ws3.title = u"你好"                             # 设定sheet名字 必须是Unicode
ws4.sheet_properties.tabColor = "1072BA"        #设定sheet的标签的背景颜色

# 选择表（sheet）
ws3 = wb["New Title"]                   # sheet 名称可以作为 key 进行索引
ws4 = wb.get_sheet_by_name("New Title") #
# ws is ws3 is ws4 --> True

# 查看表名
print(wb.sheetnames)                    # 显示所有表名
# ['Sheet2', 'New Title',  'Sheet1']
#
# for sheet in  wb:                     # 遍历所有表
# 	print(sheet.title)

# 复制一个表
wb["New Title" ]["A1"]="zeke"
source = wb["New Title" ]
target = wb.copy_worksheet(source)


# 存储数据
ws['A1'] = 42           # 方式一：数据可以直接分配到单元格中(可以输入公式)
ws.append([1, 2, 3])    # 方式二：可以附加行，从第一列开始附加(从最下方空白处，最左开始)(可以输入多行)
ws['A3'] = datetime.datetime.now().strftime("%Y-%m-%d")     # 方式三：Python 类型会被自动转换

print(rows[0][0].value) # 获取第一行第一列的单元格对象的值

# 操作单列
print(ws1["A"])
for cell in ws1["A"]:
    print cell.value

# 操作多列,获取每一个值
print(ws1["A:C"])
for column in ws1["A:C"]:
    for cell in column:
        print cell.value

# 操作多行
row_range = ws1[1:3]
print(row_range)
for row in row_range:
    for cell in row:
        print(cell.value)

wb = load_workbook('e:\\sample.xlsx')
ws=wb.active
cols=[]
cols = []
for col in ws.iter_cols():
    cols.append(col)

print(cols)                 # 所有列
print(cols[0])              # 获取第一列
print(cols[0][0])           # 获取第一列的第一行的单元格对象
print(cols[0][0].value)     # 获取第一列的第一行的值

# 访问单元格
c = ws['A4']                                    # 方法一
d = ws.cell(row=4, column=2, value=10)          # 方法二：row 行；column 列
for i in  range(1,101):                         # 方法三：只要访问就创建
        for j in range(1,101):
           ws.cell(row=i, column=j)

# 多单元格访问
cell_range = ws['A1':'C2']                      # 通过切片
colC = ws['C']                                  # 通过行(列)
col_range = ws['C:D']
row10 = ws[10]
row_range = ws[5:10]

# 通过指定范围
for row in  ws.iter_rows(min_row=1, max_row=2, max_col=3):
   for cell in  row:
       print(cell)
# <Cell Sheet1.A1>
# <Cell Sheet1.B1>
# <Cell Sheet1.C1>
# <Cell Sheet1.A2>
# <Cell Sheet1.B2>
# <Cell Sheet1.C2>

# 保存数据
wb.save('文件名称.xlsx')

# 改变 sheet 标签按钮颜色
ws.sheet_properties.tabColor = "1072BA"

# 获取最大行，最大列
print(sheet.max_row)
print(sheet.max_column)

# 获取每一行，每一列
# sheet.rows为生成器, 里面是每一行的数据，每一行又由一个tuple包裹
# sheet.columns类似，不过里面是每个tuple是每一列的单元格

for row in sheet.rows:          # 因为按行，所以返回A1, B1, C1这样的顺序
    for cell in row:
        print(cell.value)

for column in sheet.columns:    # A1, A2, A3这样的顺序
    for cell in column:
        print(cell.value)

# 根据数字得到字母，根据字母得到数字
from openpyxl.utils import get_column_letter, column_index_from_string
# 根据列的数字返回字母
print(get_column_letter(2))  # B
# 根据字母返回列的数字
print(column_index_from_string('D'))  # 4

# 删除工作表
wb.remove(sheet)        # 方式一
del wb[sheet]           # 方式二

# 设置单元格风格
from openpyxl.styles import Font, colors, Alignment
# 等线24号，加粗斜体，字体颜色红色。直接使用cell的font属性
style = Font(name='等线', size=24, italic=True, color=colors.RED, bold=True)
sheet['A1'].font = style

# 设置B1中的数据垂直居中和水平居中
sheet['B1'].alignment = Alignment(horizontal='center', vertical='center')

# 设置行高和列宽
sheet.row_dimensions[2].height = 40         # 第2行行高
sheet.column_dimensions['C'].width = 30     # C列列宽


# 注：
# object·font(字体类)：             字号、字体颜色、下划线等
# object·fill(填充类)：             颜色等
# object·border(边框类)：           设置单元格边框
# object·alignment(位置类)：        对齐方式
# object·number_format(格式类)：    数据格式
# object·protection(保护类)：       写保护


# 合并和拆分单元格
# 所谓合并单元格，即以合并区域的左上角的那个单元格为基准，覆盖其他单元格使之称为一个大的单元格
# 相反，拆分单元格后将这个大单元格的值返回到原来的左上角位置
# 合并单元格， 往左上角写入数据即可
sheet.merge_cells('B1:G1')                  # 合并一行中的几个单元格
sheet.merge_cells('A1:C3')                  # 合并一个矩形区域中的单元格

# 合并后只可以往左上角写入数据，也就是区间中:左边的坐标。
# 如果这些要合并的单元格都有数据，只会保留左上角的数据，其他则丢弃。
# 换句话说若合并前不是在左上角写入数据，合并后单元格中不会有数据。
# 以下是拆分单元格的代码。拆分后，值回到A1位置
sheet.unmerge_cells('A1:C3')


# 使用公式
wb = load_workbook('e:\\sample.xlsx')
ws1=wb.active
ws1["A1"]=1
ws1["A2"]=2
ws1["A3"]=3
ws1["A4"] = "=SUM(1, 1)"
ws1["A5"] = "=SUM(A1:A3)"
print(ws1["A4"].value)      # 打印的是公式内容，不是公式计算后的值,程序无法取到计算后的值
print(ws1["A5"].value)      # 打印的是公式内容，不是公式计算后的值,程序无法取到计算后的值


# 画一个柱状图
# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, Series

wb = load_workbook('e:\\sample.xlsx')
ws1=wb.active

wb = Workbook()
ws = wb.active
for i in range(10):
    ws.append([i])

values = Reference(ws, min_col=1, min_row=1, max_col=1, max_row=10)
chart = BarChart()
chart.add_data(values)
ws.add_chart(chart, "E15")

# Save the file
wb.save("e:\\sample.xlsx")


#画一个饼图
# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.chart import (PieChart , ProjectedPieChart, Reference)
from openpyxl.chart.series import DataPoint

data = [
    ['Pie', 'Sold'],
    ['Apple', 50],
    ['Cherry', 30],
    ['Pumpkin', 10],
    ['Chocolate', 40],
]

wb = Workbook()
ws = wb.active

for row in data:
    ws.append(row)

pie = PieChart()
labels = Reference(ws, min_col=1, min_row=2, max_row=5)
data = Reference(ws, min_col=2, min_row=1, max_row=5)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.title = "Pies sold by category"

# Cut the first slice out of the pie
slice = DataPoint(idx=0, explosion=20)
pie.series[0].data_points = [slice]

ws.add_chart(pie, "D1")

ws = wb.create_sheet(title="Projection")

data = [
    ['Page', 'Views'],
    ['Search', 95],
    ['Products', 4],
    ['Offers', 0.5],
    ['Sales', 0.5],
]

for row in data:
    ws.append(row)

projected_pie = ProjectedPieChart()
projected_pie.type = "pie"
projected_pie.splitType = "val" # split by value
labels = Reference(ws, min_col=1, min_row=2, max_row=5)
data = Reference(ws, min_col=2, min_row=1, max_row=5)
projected_pie.add_data(data, titles_from_data=True)
projected_pie.set_categories(labels)

ws.add_chart(projected_pie, "A10")

from copy import deepcopy
projected_bar = deepcopy(projected_pie)
projected_bar.type = "bar"
projected_bar.splitType = 'pos' # split by position

ws.add_chart(projected_bar, "A27")

# Save the file
wb.save("e:\\sample.xlsx")

# 3D条形图
from openpyxl import Workbook
from openpyxl.chart import (
    Reference,
    Series,
    BarChart3D,
)

wb = Workbook()
ws = wb.active

rows = [
    (None, 2013, 2014),
    ("Apples", 5, 4),
    ("Oranges", 6, 2),
    ("Pears", 8, 3)
]

for row in rows:
    ws.append(row)

data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=4)
titles = Reference(ws, min_col=1, min_row=2, max_row=4)
chart = BarChart3D()
chart.title = "3D Bar Chart"
chart.add_data(data=data, titles_from_data=True)
chart.set_categories(titles)

ws.add_chart(chart, "E5")
wb.save("bar3d.xlsx")