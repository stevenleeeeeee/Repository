#!/usr/bin/env python
#coding=utf-8

#---------------------------------------------------------
#需要先下载处理Excel的python模块：openpyxl
#本脚本用于将项目特定SHEET页下的：【IP】、【目录】列，并按行为基础，将每行结合日志的20个监控关键字输出为20行到EXCEL中
#USE:
#	python demo.py <EXCEL源文件> <第X个SHEET页> <第X列> <第X列>
#
#
#注：可自动略过EXCEL源文件中的空白行
#---------------------------------------------------------

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import sys
import time 

HOSTS=list()
MODULE=list()

#读取
INPUT_EXCEL=sys.argv[1]
EXCEL_SHEET=sys.argv[2]
EXCEL_COLUMN_IP=sys.argv[3]
EXCEL_COLUMN_MODULE=sys.argv[4]

#获取输入文件及内部的SHEET页
rb = load_workbook(filename=str(INPUT_EXCEL))
rs=rb.get_sheet_by_name(rb.sheetnames[int(EXCEL_SHEET)])

#获取所有行
rows = rs.rows

#迭代所有行的指定列
START_NUM=1
START_ROW=2
for row in rows:
	if START_NUM >= START_ROW:				#跳过EXCEL中第一行标题后再输入
		line = [col.value for col in row]
		HOSTS.append(line[int(EXCEL_COLUMN_IP)])
		MODULE.append(line[int(EXCEL_COLUMN_MODULE)])	
	START_NUM+=1

#转换后写入
wb = Workbook()
ws = wb.active

#每个模块的LOG检查关键字
CHECK_STR=[
"OnestBusiFileUploadFailed:fileType=P",
"OnestBusiFileDownloadFailed:fileType=P",
"OnestBusiFileUploadFailed:fileType=W",
"OnestBusiFileDownloadFailed:fileType=W",
"OnestBusiFileUploadFailed:fileType=V",
"OnestBusiFileDownloadFailed:fileType=V",
"OnestGztFileUploadFailed:fileType=GZT",
"OnestGztFileDownloadFailed:fileType=GZT",
"RnfsBusiFileUploadFailed",
"RnfsBusiFileDownloadFailed",
"RnfsGztFileUploadFailed:fileType=BusiFile",
"RnfsGztFileDownloadFailed:fileType=BusiFile",
"sendMqMsghasFaild",
"reqLinkfacefailedCode",
"initJvmCacheDataFailed",
"initRedisCacheDataFailed",
"Send message to vertica TOPIC error",
"BusinessFlowController GeneralException",
"BusinessFlowController Exception",
"BusinessFlowController error"
]

RECORD=0
x=1
for a in xrange(len(HOSTS)):
	for b in xrange(len(CHECK_STR)):
		if len(str(HOSTS[a])) >= 5 and len(str(MODULE[a])) >= 5:
			ws.cell(row=int(x),column=1).value=HOSTS[a]		#根据关键字个数输出N个相同的主机
			ws.cell(row=int(x),column=2).value=MODULE[a]		#根据关键字个数输出N个相同的模块地址
			ws.cell(row=int(x),column=3).value=CHECK_STR[b]		#输出关键字
			x+=1
		
if __name__ == '__main__':
	OUT_FILENAME=time.strftime("%Y-%m-%d", time.localtime()) 
	wb.save(filename=str(OUT_FILENAME)+".xlsx")
	print("--"*30)
	print("执行成功!... 请查看输出文件: "+str(OUT_FILENAME)+".xlsx")
	print("共转换输出了 "+str(x-1)+" 行记录")
